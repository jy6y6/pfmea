import streamlit as st
import pandas as pd
import requests
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

# ===================== 核心配置（已内置完成，无需修改）=====================
# 内置API密钥，开箱即用
API_KEY = "7abbafd6-4d6e-4dad-9172-ea2d165c7a44"
API_ENDPOINT = "https://api.doubao.com/v1/chat/completions"
# 系统基础配置
SYSTEM_NAME = "电池包/充电器PFMEA智能生成系统"
STANDARD = "AIAG-VDA FMEA 第一版 | IATF16949:2016"

# ===================== 1. 全工序专业本地标准库（符合审核要求）=====================
# 电池包装配工序标准库
BATTERY_PROCESS_LIB = {
    "电芯来料检验": [
        {
            "失效模式": "电芯外观尺寸超差",
            "失效后果": "电芯无法装入模组壳体，导致装配中断，产生返工成本",
            "失效原因": "来料尺寸公差不符合图纸要求，检验量具未定期校准",
            "预防措施": "制定电芯来料检验规范，每批次抽取样件全尺寸检测，量具定期校准并记录",
            "探测措施": "首件全尺寸检验，巡检按AQL抽样标准检测，超差件隔离标识",
            "严重度S": 6,
            "频度O": 3,
            "探测度D": 4,
            "AP等级": "中"
        },
        {
            "失效模式": "电芯电压/内阻异常",
            "失效后果": "模组充放电异常，循环寿命衰减过快，严重时引发热失控风险",
            "失效原因": "电芯生产过程工艺异常，来料存储环境温湿度不符合要求",
            "预防措施": "每批次电芯进行电压、内阻全检，存储环境温湿度24小时监控记录",
            "探测措施": "自动化检测设备100%全检，异常数据自动报警隔离，数据可追溯",
            "严重度S": 9,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "电芯外观磕碰、掉角、铝塑膜破损",
            "失效后果": "电芯漏液、绝缘失效，引发短路、热失控安全风险",
            "失效原因": "来料运输过程防护不当，检验过程操作不规范造成二次损伤",
            "预防措施": "来料包装规范验证，检验人员操作培训考核，制定外观缺陷判定标准",
            "探测措施": "100%外观目视+放大镜检测，缺陷件隔离报废，记录追溯",
            "严重度S": 8,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "高"
        }
    ],
    "模组堆叠装配": [
        {
            "失效模式": "电芯堆叠顺序错误、极性反向",
            "失效后果": "模组电路连接错误，充放电功能失效，严重时引发短路烧毁",
            "失效原因": "作业人员未按SOP操作，防错装置失效，首件检验未执行",
            "预防措施": "制定极性防错SOP，安装极性视觉防错装置，作业人员岗前培训考核",
            "探测措施": "首件极性全检，过程中视觉设备100%检测，异常自动停机报警",
            "严重度S": 9,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "电芯堆叠间隙超差",
            "失效后果": "模组尺寸超差无法装入PACK箱体，热管理路径失效，电芯散热不均",
            "失效原因": "堆叠工装定位偏差，缓冲泡棉厚度不符合要求，作业人员操作不当",
            "预防措施": "堆叠工装定期校准维护，泡棉来料全尺寸检验，制定堆叠作业标准",
            "探测措施": "首件尺寸全检，过程中用塞尺巡检，超差件返工调整",
            "严重度S": 6,
            "频度O": 3,
            "探测度D": 4,
            "AP等级": "中"
        },
        {
            "失效模式": "电芯捆绑力不足/过大",
            "失效后果": "捆绑力不足导致电芯膨胀位移，过大导致电芯壳体变形损伤",
            "失效原因": "捆绑设备扭矩未校准，作业人员未按标准参数操作",
            "预防措施": "捆绑设备每日班前校准扭矩，制定标准扭矩参数，作业人员培训考核",
            "探测措施": "首件扭矩检测，过程中每小时巡检，扭矩数据记录存档",
            "严重度S": 7,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        }
    ],
    "母线排激光焊接": [
        {
            "失效模式": "焊接虚焊、假焊、脱焊",
            "失效后果": "回路电阻过大，充放电过程发热严重，引发绝缘失效、烧毁风险",
            "失效原因": "激光功率参数不符，焊接表面有油污/氧化层，工装定位偏差",
            "预防措施": "制定焊接参数规范，焊接前表面清洁处理，工装定期校准定位精度",
            "探测措施": "首件焊接拉力测试+金相切片分析，过程中100%视觉检测，拉力测试定期抽检",
            "严重度S": 8,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "高"
        },
        {
            "失效模式": "焊接焊穿、炸点",
            "失效后果": "电芯极耳破损漏液，母线排结构强度不足，引发短路安全风险",
            "失效原因": "激光功率过大，焊接速度过慢，聚焦高度偏差",
            "预防措施": "焊接参数DOE验证锁定，每日班前首件参数确认，设备定期维护校准",
            "探测措施": "100%视觉检测焊道外观，首件切片分析，缺陷件隔离报废",
            "严重度S": 9,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "焊接飞溅残留",
            "失效后果": "金属飞溅导致高压回路短路，绝缘性能下降，引发安全事故",
            "失效原因": "焊接保护气体流量不足，焊接区域防护不到位，清洁工序遗漏",
            "预防措施": "焊接区域加装防护挡板，保护气体流量参数锁定，制定焊后清洁SOP",
            "探测措施": "焊后100%高压吹气清洁+目视检测，绝缘耐压测试验证",
            "严重度S": 8,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "高"
        }
    ],
    "模组绝缘耐压测试": [
        {
            "失效模式": "绝缘耐压测试不通过",
            "失效后果": "模组高压绝缘失效，装车后引发漏电、触电、短路烧毁风险",
            "失效原因": "焊接飞溅残留，绝缘片破损，母线排与壳体间隙不足",
            "预防措施": "焊后全工序清洁管控，绝缘片安装防错设计，装配过程绝缘防护",
            "探测措施": "100%绝缘耐压测试，测试设备定期校准，不合格件隔离返工",
            "严重度S": 9,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "测试参数设置错误",
            "失效后果": "不合格品流出，存在严重安全隐患，客户审核不符合项",
            "失效原因": "作业人员未按标准设置参数，测试程序未锁定，无防错验证",
            "预防措施": "测试程序加密锁定，制定标准测试参数规范，作业人员岗前培训考核",
            "探测措施": "首件参数确认，每批次首件测试标准样件验证，测试记录存档追溯",
            "严重度S": 8,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "高"
        },
        {
            "失效模式": "测试探针接触不良",
            "失效后果": "测试数据失真，不合格品误判流出，引发客户投诉",
            "失效原因": "探针磨损、氧化，测试工装定位偏差，探针压力不足",
            "预防措施": "制定探针定期更换维护计划，工装每日班前校准，探针压力定期检测",
            "探测措施": "每日班前用标准样件验证测试设备，异常探针立即更换",
            "严重度S": 7,
            "频度O": 3,
            "探测度D": 4,
            "AP等级": "中"
        }
    ],
    "PACK箱体装配": [
        {
            "失效模式": "箱体密封性能不达标",
            "失效后果": "水汽、粉尘进入箱体内部，引发绝缘失效、元器件腐蚀，IP等级不达标",
            "失效原因": "密封胶条选型不符，涂胶路径/厚度不符合要求，螺栓紧固扭矩不均",
            "预防措施": "密封胶条来料检验，涂胶设备参数锁定，螺栓紧固扭矩标准制定",
            "探测措施": "100%气密性测试，测试设备定期校准，不合格件隔离返工",
            "严重度S": 8,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "高"
        },
        {
            "失效模式": "模组安装定位偏差",
            "失效后果": "高压件连接错位，箱体无法正常合盖，高压安全间隙不足",
            "失效原因": "安装工装定位偏差，箱体安装孔位尺寸超差，作业人员操作不当",
            "预防措施": "安装工装定期校准，箱体来料孔位全检，制定模组安装SOP",
            "探测措施": "首件安装尺寸全检，过程中巡检定位精度，超差件返工调整",
            "严重度S": 6,
            "频度O": 3,
            "探测度D": 4,
            "AP等级": "中"
        },
        {
            "失效模式": "箱体紧固件漏装、扭矩不符",
            "失效后果": "箱体连接强度不足，车辆行驶过程中松动脱落，引发高压安全风险",
            "失效原因": "作业人员漏打螺栓，扭矩扳手未校准，无防错计数装置",
            "预防措施": "使用带计数功能的电动扭矩扳手，扭矩每日班前校准，制定紧固顺序SOP",
            "探测措施": "首件扭矩全检，过程中划线防错，每小时巡检扭矩值，记录存档",
            "严重度S": 8,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "高"
        }
    ],
    "高压线束装配": [
        {
            "失效模式": "高压插件漏插、插合不到位",
            "失效后果": "高压回路断路，充放电功能失效，车辆行驶中断电，引发安全事故",
            "失效原因": "作业人员未按SOP操作，插件卡扣损坏，无到位防错装置",
            "预防措施": "制定插件装配SOP，插件来料检验，作业人员岗前培训考核",
            "探测措施": "插件插合后100%卡扣到位目视检查，高压导通测试验证，不合格件返工",
            "严重度S": 9,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "高"
        },
        {
            "失效模式": "高压线束固定不当、磨损",
            "失效后果": "线束绝缘层磨损破损，引发高压短路、漏电，车辆起火风险",
            "失效原因": "线束走向不符合图纸，固定卡扣漏装/选型错误，与尖锐边缘无防护",
            "预防措施": "制定线束走向标准，卡扣选型与安装数量锁定，尖锐边缘加装防护套",
            "探测措施": "首件线束走向全检，过程中100%目视检查，高压绝缘测试验证",
            "严重度S": 9,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "高"
        },
        {
            "失效模式": "线束端子压接不良",
            "失效后果": "回路电阻过大，工作过程发热严重，引发端子烧毁、绝缘失效",
            "失效原因": "压接模具选型错误，压接设备参数未校准，端子/线材来料不符",
            "预防措施": "压接模具与线材匹配验证，设备参数每日校准，制定压接作业标准",
            "探测措施": "首件压接拉力测试+剖面分析，过程中每小时抽检拉力，记录存档",
            "严重度S": 8,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "高"
        }
    ],
    "BMS控制器装配与接线": [
        {
            "失效模式": "BMS采样线接错、漏接",
            "失效后果": "电芯电压采集异常，BMS控制逻辑失效，过充过放引发热失控风险",
            "失效原因": "作业人员未按接线图操作，采样线编号错误，无防错校验",
            "预防措施": "采样线编号与接口一一对应，制定接线SOP，作业人员岗前培训考核",
            "探测措施": "接线后100%导通测试，BMS上电后采集数据校验，异常自动报警",
            "严重度S": 9,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "BMS控制器固定不当",
            "失效后果": "车辆行驶过程中控制器松动脱落，接线断裂，电池管理系统完全失效",
            "失效原因": "紧固件扭矩不符，漏装防松垫圈，安装孔位错位",
            "预防措施": "使用定扭矩工具紧固，防松垫圈强制安装，安装孔位来料检验",
            "探测措施": "首件扭矩全检，过程中划线防错巡检，安装后晃动验证牢固度",
            "严重度S": 9,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "高"
        },
        {
            "失效模式": "BMS低压供电异常",
            "失效后果": "控制器无法正常上电，电池管理系统失效，无法监控电池状态",
            "失效原因": "低压保险选型错误，接线端子虚接，供电线路短路",
            "预防措施": "保险规格图纸锁定，端子压接标准制定，线路绝缘防护管控",
            "探测措施": "上电前低压回路导通+绝缘测试，上电后电压全检，异常隔离返工",
            "严重度S": 8,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "高"
        }
    ],
    "热管理系统装配": [
        {
            "失效模式": "冷却管路漏液",
            "失效后果": "冷却液泄漏导致热管理失效，电芯温度失控，绝缘短路风险",
            "失效原因": "管路接头紧固扭矩不符，密封圈破损漏装，管路老化开裂",
            "预防措施": "接头扭矩标准锁定，密封圈安装防错SOP，管路来料耐压测试",
            "探测措施": "装配后100%保压测试，测试数据记录存档，不合格件隔离返工",
            "严重度S": 8,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "高"
        },
        {
            "失效模式": "导热垫贴合不良",
            "失效后果": "电芯与冷却板导热路径失效，电芯散热不均，温差过大，循环寿命衰减",
            "失效原因": "导热垫厚度选型错误，表面保护膜未撕除，安装压力不均",
            "预防措施": "导热垫来料厚度全检，制定安装SOP，保护膜撕除双人复核",
            "探测措施": "安装后贴合间隙塞尺检测，热循环测试验证温差，记录存档",
            "严重度S": 7,
            "频度O": 3,
            "探测度D": 4,
            "AP等级": "中"
        },
        {
            "失效模式": "冷却泵/风扇功能异常",
            "失效后果": "冷却液循环/风冷散热失效，电池包过热，充放电功率受限",
            "失效原因": "泵/风扇接线错误，来料功能不良，装配过程磕碰损伤",
            "预防措施": "来料功能全检，装配过程防护管控，接线标准SOP制定",
            "探测措施": "装配后100%功能测试，转速、流量、噪音全检，异常件返工",
            "严重度S": 7,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        }
    ],
    "PACK总成绝缘耐压测试": [
        {
            "失效模式": "绝缘耐压测试不通过",
            "失效后果": "高压系统绝缘失效，装车后引发漏电、触电、短路起火，严重安全事故",
            "失效原因": "高压件与壳体间隙不足，金属异物残留，绝缘件破损，密封失效进水",
            "预防措施": "装配过程全流程异物管控，绝缘件安装防护，高压间隙设计验证",
            "探测措施": "100%绝缘耐压测试，测试设备定期校准，不合格件100%隔离返工",
            "严重度S": 10,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "测试过程高压放电伤人",
            "失效后果": "作业人员触电伤亡，安全生产事故",
            "失效原因": "作业人员未按安全规范操作，测试区域无安全防护，无高压放电流程",
            "预防措施": "制定高压安全操作规范，作业人员持证上岗，测试区域加装安全联锁",
            "探测措施": "测试前安全条件确认，双人复核操作，测试后强制放电流程",
            "严重度S": 10,
            "频度O": 1,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "测试记录不完整、可追溯性差",
            "失效后果": "客户审核不符合项，不合格品无法追溯，质量问题无法闭环",
            "失效原因": "测试数据未自动存档，作业人员记录不规范，无数据备份机制",
            "预防措施": "测试设备与MES系统对接，数据自动上传存档，制定记录管理规范",
            "探测措施": "每批次测试记录复核，审核前全量检查，记录存档期限符合法规要求",
            "严重度S": 5,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "中"
        }
    ],
    "PACK总成充放电测试": [
        {
            "失效模式": "充放电容量不达标",
            "失效后果": "电池包续航里程不足，客户投诉，产品退货",
            "失效原因": "电芯一致性差，BMSSOC校准错误，充放电参数设置不符",
            "预防措施": "电芯一致性分选管控，BMSSOC校准规范制定，测试程序参数锁定",
            "探测措施": "100%充放电容量测试，数据自动记录，不合格件隔离返工",
            "严重度S": 7,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "中"
        },
        {
            "失效模式": "充放电过程温差过大",
            "失效后果": "电芯循环寿命衰减过快，严重时引发热失控风险",
            "失效原因": "热管理系统功能异常，电芯一致性差，充放电倍率过大",
            "预防措施": "热管理系统装配前功能测试，电芯一致性分选，充放电参数锁定",
            "探测措施": "测试过程全时段温度监控，温差超标自动报警，数据记录存档",
            "严重度S": 8,
            "频度O": 3,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "BMS保护功能异常",
            "失效后果": "过充、过放、过流、短路保护失效，引发电池包热失控、起火安全事故",
            "失效原因": "BMS固件版本错误，保护参数设置不符，硬件采样电路故障",
            "预防措施": "BMS固件版本锁定，保护参数按图纸设置，来料硬件功能全检",
            "探测措施": "测试过程中保护功能全项验证，异常自动报警，不合格件报废",
            "严重度S": 10,
            "频度O": 1,
            "探测度D": 2,
            "AP等级": "高"
        }
    ],
    "PACK外观与终检": [
        {
            "失效模式": "外观标识错误、漏贴",
            "失效后果": "产品型号混淆，客户安装错误，合规性不达标，审核不符合项",
            "失效原因": "标识打印错误，作业人员漏贴，标识型号与产品不匹配",
            "预防措施": "标识打印与产品订单绑定防错，制定标识粘贴SOP，作业人员培训",
            "探测措施": "100%外观目视检查，双人复核标识型号，不合格件返工",
            "严重度S": 5,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "中"
        },
        {
            "失效模式": "箱体外观磕碰、划伤、锈蚀",
            "失效后果": "客户投诉，产品外观不合格，严重时影响箱体防护性能",
            "失效原因": "装配过程操作不当，运输过程防护不足，来料表面处理不良",
            "预防措施": "装配过程箱体加装防护套，制定外观缺陷判定标准，来料外观全检",
            "探测措施": "100%外观目视检查，缺陷件隔离返工，记录存档",
            "严重度S": 4,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "低"
        },
        {
            "失效模式": "附件漏装、错装",
            "失效后果": "客户无法正常安装使用，客户投诉，产品退货",
            "失效原因": "作业人员未按装箱清单操作，无防错计数装置",
            "预防措施": "制定装箱清单SOP，附件按单定量配送，双人复核装箱",
            "探测措施": "100%装箱清单核对，终检全项检查，记录存档",
            "严重度S": 5,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        }
    ],
    "产品包装入库": [
        {
            "失效模式": "包装防护不当",
            "失效后果": "运输过程中产品磕碰、损伤，客户投诉，产品返工",
            "失效原因": "包装材料选型错误，固定方式不当，防水防潮措施不足",
            "预防措施": "包装方案运输验证，包装材料来料检验，制定包装作业SOP",
            "探测措施": "100%包装外观检查，首件包装防护验证，不合格件返工",
            "严重度S": 4,
            "频度O": 3,
            "探测度D": 4,
            "AP等级": "低"
        },
        {
            "失效模式": "产品批次信息错误",
            "失效后果": "产品批次混乱，质量问题无法追溯，客户审核不符合项",
            "失效原因": "外箱标签打印错误，产品与外箱批次不匹配，扫码防错失效",
            "预防措施": "标签打印与产品批次绑定，扫码防错系统对接MES，作业人员培训",
            "探测措施": "100%扫码核对批次信息，双人复核，记录存档",
            "严重度S": 6,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        },
        {
            "失效模式": "仓储环境不符合要求",
            "失效后果": "产品电芯自放电过大，性能衰减，金属件锈蚀，产品失效",
            "失效原因": "仓储温湿度未管控，无定期巡检记录，产品堆叠方式不当",
            "预防措施": "仓储环境温湿度24小时监控，制定仓储管理规范，定期巡检",
            "探测措施": "每日温湿度记录检查，定期库存产品电压抽检，异常隔离处理",
            "严重度S": 7,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        }
    ]
}

# 充电器装配工序标准库
CHARGER_PROCESS_LIB = {
    "PCB来料检验": [
        {
            "失效模式": "PCB板尺寸、孔位超差",
            "失效后果": "PCB无法装入壳体，元器件安装错位，装配中断",
            "失效原因": "PCB生产制程偏差，来料检验规范未执行，量具未校准",
            "预防措施": "制定PCB来料检验规范，每批次首件全尺寸检测，量具定期校准",
            "探测措施": "首件全尺寸检验，巡检按AQL抽样，超差件隔离返工",
            "严重度S": 5,
            "频度O": 3,
            "探测度D": 4,
            "AP等级": "中"
        },
        {
            "失效模式": "PCB线路开路、短路",
            "失效后果": "电路板功能失效，产品无法正常工作，严重时引发烧毁",
            "失效原因": "PCB生产蚀刻工艺异常，来料运输过程静电损伤",
            "预防措施": "PCB来料100%通断测试，运输过程防静电包装，检验过程防静电管控",
            "探测措施": "自动化测试设备100%通断检测，异常板隔离报废，数据追溯",
            "严重度S": 8,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "PCB表面镀层不良、氧化",
            "失效后果": "元器件焊接虚焊，接触电阻过大，产品寿命衰减",
            "失效原因": "PCB表面处理工艺异常，来料存储环境温湿度不符合要求",
            "预防措施": "PCB镀层来料检验，存储环境温湿度管控，制定存储有效期规范",
            "探测措施": "100%外观目视检测，首件可焊性测试，不良品隔离",
            "严重度S": 6,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "中"
        }
    ],
    "SMT贴片焊接": [
        {
            "失效模式": "元器件贴装偏移、错件、漏件",
            "失效后果": "电路功能失效，产品无法正常工作，批量返工成本",
            "失效原因": "贴片机程序错误，元器件料盘上错，吸嘴磨损定位偏差",
            "预防措施": "贴片机程序首件验证，上料双人复核，设备定期维护保养",
            "探测措施": "首件全项核对，SPI锡膏检测，AOI光学100%检测，异常报警",
            "严重度S": 7,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "焊接虚焊、假焊、连锡",
            "失效后果": "回路接触不良，产品工作异常，短路烧毁，寿命衰减",
            "失效原因": "锡膏印刷不良，回流焊温度曲线不符，元器件引脚氧化",
            "预防措施": "锡膏印刷参数锁定，回流焊温度曲线DOE验证，元器件存储管控",
            "探测措施": "SPI锡膏检测，AOI光学检测，首件X-Ray检测，定期切片分析",
            "严重度S": 8,
            "频度O": 3,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "元器件静电损伤",
            "失效后果": "元器件隐性损伤，产品早期失效，客户投诉退货",
            "失效原因": "生产过程无防静电管控，作业人员未按规范操作",
            "预防措施": "车间全区域防静电管控，作业人员防静电装备穿戴，岗前培训考核",
            "探测措施": "每日防静电设备点检，产品功能全检，异常品失效分析",
            "严重度S": 7,
            "频度O": 2,
            "探测度D": 4,
            "AP等级": "中"
        }
    ],
    "DIP插件焊接": [
        {
            "失效模式": "插件引脚漏插、插反、错插",
            "失效后果": "电路功能失效，极性反接导致元器件烧毁，产品报废",
            "失效原因": "作业人员未按SOP操作，元器件极性标识不清，无防错措施",
            "预防措施": "制定插件作业SOP，元器件极性标识清晰，作业人员岗前培训考核",
            "探测措施": "插件后目视检查，首件全项核对，焊后功能测试验证",
            "严重度S": 8,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "高"
        },
        {
            "失效模式": "焊锡不良、虚焊、堆锡、拉尖",
            "失效后果": "接触电阻过大，产品发热异常，短路烧毁，功能失效",
            "失效原因": "焊锡温度参数不符，作业人员操作不规范，助焊剂选型错误",
            "预防措施": "焊锡机参数锁定，制定焊接作业标准，作业人员培训考核持证上岗",
            "探测措施": "焊后100%目视检查，首件电气性能测试，不良品隔离返工",
            "严重度S": 7,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "中"
        },
        {
            "失效模式": "PCB焊盘脱落",
            "失效后果": "电路断路，产品功能失效，PCB板报废",
            "失效原因": "焊接温度过高，焊接时间过长，PCB焊盘附着力不足",
            "预防措施": "焊接温度与时间参数锁定，PCB来料焊盘附着力测试，作业规范制定",
            "探测措施": "焊后目视检查，首件拉力测试，不良品隔离报废",
            "严重度S": 6,
            "频度O": 2,
            "探测度D": 4,
            "AP等级": "中"
        }
    ],
    "PCBA功能测试": [
        {
            "失效模式": "输入输出电压异常",
            "失效后果": "产品无法正常供电，烧毁负载设备，引发客户投诉",
            "失效原因": "电源芯片参数不符，反馈电路元器件不良，测试设备未校准",
            "预防措施": "元器件来料全检，测试程序参数锁定，测试设备定期校准",
            "探测措施": "100%自动化电压测试，异常数据自动报警隔离，记录存档",
            "严重度S": 8,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "保护功能失效",
            "失效后果": "过压、过流、短路、过热保护失效，引发产品烧毁、火灾安全风险",
            "失效原因": "保护电路元器件不良，固件参数设置错误，测试项目遗漏",
            "预防措施": "保护电路设计验证，固件参数锁定，制定全项测试规范",
            "探测措施": "100%保护功能全项测试，异常品隔离报废，测试数据可追溯",
            "严重度S": 10,
            "频度O": 1,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "空载功耗、效率不达标",
            "失效后果": "产品能耗超标，不符合能效法规要求，无法通过认证，客户退货",
            "失效原因": "变压器设计不良，元器件损耗过大，测试环境不符",
            "预防措施": "电路设计仿真验证，元器件选型锁定，测试环境温湿度管控",
            "探测措施": "每批次首件全性能测试，过程中抽检，不合格品隔离返工",
            "严重度S": 6,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        }
    ],
    "壳体装配": [
        {
            "失效模式": "壳体上下盖合缝间隙超差",
            "失效后果": "产品外观不良，客户投诉，IP防护等级不达标",
            "失效原因": "壳体注塑尺寸超差，装配卡扣配合不良，螺丝紧固扭矩不均",
            "预防措施": "壳体来料尺寸全检，装配工装定位校准，螺丝扭矩标准锁定",
            "探测措施": "装配后100%目视+塞尺检测，超差件返工调整",
            "严重度S": 4,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "低"
        },
        {
            "失效模式": "PCBA固定不当、松动",
            "失效后果": "产品运输过程中PCBA位移，引脚断裂，功能失效，客户投诉",
            "失效原因": "固定螺丝漏装、扭矩不符，定位柱尺寸超差，减震垫漏装",
            "预防措施": "使用定扭矩工具，螺丝计数防错，定位柱来料检验，减震垫安装SOP",
            "探测措施": "首件扭矩全检，装配后晃动验证，100%目视检查",
            "严重度S": 7,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        },
        {
            "失效模式": "壳体卡扣断裂",
            "失效后果": "壳体合盖不严，外观不良，IP防护失效，产品报废",
            "失效原因": "壳体注塑材料脆性大，作业人员装配操作不当，卡扣设计强度不足",
            "预防措施": "壳体材料来料验证，制定装配操作SOP，作业人员岗前培训",
            "探测措施": "装配后100%目视检查，卡扣断裂件隔离报废",
            "严重度S": 5,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "中"
        }
    ],
    "输入输出线装配": [
        {
            "失效模式": "线材端子压接不良",
            "失效后果": "回路电阻过大，工作发热严重，端子烧毁，产品功能失效",
            "失效原因": "压接模具选型错误，压接设备参数未校准，线材/端子来料不符",
            "预防措施": "压接模具与线材匹配验证，设备参数每日校准，制定压接作业标准",
            "探测措施": "首件拉力测试+剖面分析，过程中每小时抽检拉力，记录存档",
            "严重度S": 8,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "高"
        },
        {
            "失效模式": "线材极性接反、错接",
            "失效后果": "输出极性错误，烧毁客户负载设备，引发安全事故，客户巨额索赔",
            "失效原因": "作业人员未按SOP操作，线材颜色标识错误，无防错校验",
            "预防措施": "线材极性颜色标识锁定，制定接线SOP，作业人员岗前培训考核",
            "探测措施": "接线后100%极性导通测试，异常自动报警，不合格品隔离",
            "严重度S": 9,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "线材固定不当、应力释放不足",
            "失效后果": "线材拉扯过程中端子脱落，电路断路，产品功能失效，安全风险",
            "失效原因": "线卡固定扭矩不符，线材预留长度不足，无应力释放结构",
            "预防措施": "线卡扭矩标准锁定，制定线材布线规范，应力释放结构设计验证",
            "探测措施": "装配后100%拉力测试，目视检查布线，不合格品返工",
            "严重度S": 7,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "中"
        }
    ],
    "灌封与固化": [
        {
            "失效模式": "灌封胶气泡、填充不满",
            "失效后果": "产品绝缘性能下降，防水防潮失效，元器件散热不良，寿命衰减",
            "失效原因": "灌封胶脱泡不充分，灌胶路径/剂量不符，固化参数设置错误",
            "预防措施": "灌胶设备参数锁定，灌封胶脱泡工艺规范，固化温度时间参数验证",
            "探测措施": "灌胶后目视检查，首件切片分析，固化后绝缘测试验证",
            "严重度S": 6,
            "频度O": 3,
            "探测度D": 4,
            "AP等级": "中"
        },
        {
            "失效模式": "灌封胶固化不完全",
            "失效后果": "产品密封性能失效，胶体流淌，外观不良，绝缘性能不达标",
            "失效原因": "AB胶配比错误，固化温度/时间不足，胶体过期失效",
            "预防措施": "灌胶设备配比自动校准，固化参数锁定，胶体有效期管控",
            "探测措施": "固化后硬度测试，首件固化验证，不合格品隔离返工",
            "严重度S": 5,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        },
        {
            "失效模式": "灌封胶污染产品外观、接插件",
            "失效后果": "产品外观不良，接插件接触不良，功能失效，客户投诉",
            "失效原因": "灌胶过程防护不到位，胶量过大溢出，作业人员操作不当",
            "预防措施": "灌胶区域加装防护工装，胶量参数锁定，制定灌胶作业SOP",
            "探测措施": "灌胶后100%目视检查，污染件清洁返工，接插件导通测试",
            "严重度S": 4,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "低"
        }
    ],
    "成品耐压绝缘测试": [
        {
            "失效模式": "绝缘耐压测试不通过",
            "失效后果": "产品绝缘失效，引发漏电、触电、短路烧毁，严重安全事故",
            "失效原因": "灌封胶绝缘性能不足，高压回路与壳体间隙不足，金属异物残留",
            "预防措施": "灌封胶来料绝缘性能验证，装配过程异物管控，高压间隙设计验证",
            "探测措施": "100%绝缘耐压测试，测试设备定期校准，不合格品隔离报废",
            "严重度S": 10,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "漏电流超标",
            "失效后果": "产品不符合安规认证要求，无法上市销售，客户退货",
            "失效原因": "Y电容选型错误，变压器绝缘不良，PCB布线设计缺陷",
            "预防措施": "元器件选型锁定，安规设计验证，来料元器件绝缘性能全检",
            "探测措施": "100%漏电流测试，数据自动记录，不合格品隔离",
            "严重度S": 8,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "高"
        },
        {
            "失效模式": "测试接地不良",
            "失效后果": "测试数据失真，不合格品误判流出，引发安全事故",
            "失效原因": "测试接地端子氧化，接地线松动，测试工装接地不良",
            "预防措施": "测试设备每日班前接地电阻检测，工装定期维护，制定点检规范",
            "探测措施": "每日班前用标准样件验证测试设备，异常立即停机整改",
            "严重度S": 9,
            "频度O": 1,
            "探测度D": 3,
            "AP等级": "高"
        }
    ],
    "成品老化测试": [
        {
            "失效模式": "老化过程产品失效、烧毁",
            "失效后果": "产品报废，批量质量隐患，生产成本增加",
            "失效原因": "产品隐性缺陷未检出，老化参数设置错误，老化设备保护失效",
            "预防措施": "老化设备保护功能定期验证，老化参数程序锁定，产品前道全检管控",
            "探测措施": "老化过程全时段电压、电流、温度监控，异常自动断电报警",
            "严重度S": 7,
            "频度O": 2,
            "探测度D": 2,
            "AP等级": "中"
        },
        {
            "失效模式": "老化后产品性能衰减",
            "失效后果": "产品早期失效，寿命不达标，客户投诉退货",
            "失效原因": "元器件选型不符，产品设计缺陷，老化条件超出规格",
            "预防措施": "元器件可靠性验证，产品设计寿命测试，老化参数规范锁定",
            "探测措施": "老化前后全性能对比测试，衰减超标品隔离报废，数据记录存档",
            "严重度S": 6,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        },
        {
            "失效模式": "老化记录不完整、可追溯性差",
            "失效后果": "客户审核不符合项，质量问题无法追溯，无法闭环",
            "失效原因": "老化数据未自动存档，作业人员记录不规范，无数据备份",
            "预防措施": "老化设备与MES系统对接，数据自动上传存档，制定记录管理规范",
            "探测措施": "每批次老化记录复核，审核前全量检查，记录存档期限符合法规",
            "严重度S": 5,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "中"
        }
    ],
    "外观终检与包装入库": [
        {
            "失效模式": "产品外观划伤、污渍、标识错误",
            "失效后果": "客户投诉，产品外观不合格，品牌形象受损",
            "失效原因": "装配过程操作不当，标识打印错误，作业人员漏检",
            "预防措施": "产品装配过程防护管控，标识打印与订单绑定防错，制定终检规范",
            "探测措施": "100%外观目视检查，双人复核标识，不合格品隔离返工",
            "严重度S": 4,
            "频度O": 3,
            "探测度D": 3,
            "AP等级": "低"
        },
        {
            "失效模式": "产品型号、附件错装漏装",
            "失效后果": "客户无法正常使用，投诉退货，品牌信誉受损",
            "失效原因": "作业人员未按装箱清单操作，无防错计数装置",
            "预防措施": "附件按单定量配送，制定装箱SOP，双人复核装箱",
            "探测措施": "100%装箱清单核对，终检全项检查，记录存档",
            "严重度S": 5,
            "频度O": 2,
            "探测度D": 3,
            "AP等级": "中"
        },
        {
            "失效模式": "包装防护不当，运输过程产品损坏",
            "失效后果": "产品损坏，客户投诉，返工成本增加",
            "失效原因": "包装材料选型错误，固定方式不当，跌落测试未通过",
            "预防措施": "包装方案跌落验证，包装材料来料检验，制定包装作业SOP",
            "探测措施": "100%包装外观检查，首件包装防护验证，不合格品返工",
            "严重度S": 4,
            "频度O": 3,
            "探测度D": 4,
            "AP等级": "低"
        }
    ]
}

# ===================== 2. 核心工具函数 =====================
# 初始化session_state，彻底解决刷新报错
def init_session_state():
    if "user_knowledge_base" not in st.session_state:
        st.session_state.user_knowledge_base = {}  # 用户自定义知识库，格式同标准库
    if "generated_pfmea_data" not in st.session_state:
        st.session_state.generated_pfmea_data = None
    if "selected_scheme" not in st.session_state:
        st.session_state.selected_scheme = None
    if "ai_schemes" not in st.session_state:
        st.session_state.ai_schemes = None
    if "current_process" not in st.session_state:
        st.session_state.current_process = None
    if "current_product" not in st.session_state:
        st.session_state.current_product = None
    if "upload_status" not in st.session_state:
        st.session_state.upload_status = None

# AI生成核心函数（重构，解决与本地库重复问题，支持多方案生成）
def generate_pfmea_ai(process_name, product_type, scheme_count=3):
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    
    # 行业专属提示词，强制多方案、差异化、贴合场景，与本地库完全区分
    prompt = f"""
    你是专业的汽车电子行业PFMEA工程师，精通AIAG-VDA FMEA标准和IATF16949质量管理体系要求，专注于{product_type}装配制造场景。
    请针对【{process_name}】工序，生成{scheme_count}组完全不同、无重复内容的PFMEA方案，每组方案包含3-5条独立的PFMEA条目。
    严格遵守以下要求：
    1. 每组方案必须有明显差异化：分别从人、机、料、法、环、测不同维度切入，失效模式、失效后果、失效原因、预防/探测措施完全不同，禁止内容重复
    2. 所有内容必须严格贴合{product_type}装配现场的实际作业场景，禁止通用化、理论化内容，必须是可落地、可执行的现场管控措施
    3. 严格遵循AIAG-VDA FMEA标准，失效链必须完整：失效模式→失效后果→失效原因→预防措施→探测措施，逻辑闭环
    4. S/O/D评分严格符合AIAG-VDA评分标准：
       - 严重度S：1-10分，安全相关必须≥8分
       - 频度O：1-10分，有预防措施的必须≤4分
       - 探测度D：1-10分，有自动化探测的必须≤3分
    5. AP等级严格按S/O/D评分判定：高/中/低三个等级，符合IATF16949审核要求
    6. 必须返回严格的JSON格式，外层是一个数组，每个元素是一组方案，格式如下：
    [
        {{
            "方案名称": "方案1：人员操作维度管控",
            "pfmea_list": [
                {{
                    "失效模式": "xxx",
                    "失效后果": "xxx",
                    "失效原因": "xxx",
                    "预防措施": "xxx",
                    "探测措施": "xxx",
                    "严重度S": x,
                    "频度O": x,
                    "探测度D": x,
                    "AP等级": "x"
                }}
            ]
        }}
    ]
    7. 禁止返回任何JSON以外的内容，禁止注释、解释、 markdown格式，确保JSON可直接解析
    8. 禁止使用本地标准库中的重复内容，必须生成全新的、差异化的PFMEA条目
    """
    
    data = {
        "model": "doubao-pro-32k",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.7,
        "top_p": 0.9,
        "max_tokens": 4000
    }
    
    try:
        response = requests.post(API_ENDPOINT, headers=headers, json=data, timeout=30)
        response.raise_for_status()
        result = response.json()
        ai_content = result["choices"][0]["message"]["content"]
        
        # 格式容错处理，提取JSON内容
        ai_content = ai_content.strip()
        if ai_content.startswith("```json"):
            ai_content = ai_content[7:]
        if ai_content.endswith("```"):
            ai_content = ai_content[:-3]
        ai_content = ai_content.strip()
        
        # 解析JSON
        schemes = json.loads(ai_content)
        return schemes, None
    
    except Exception as e:
        # 异常兜底，仅当AI完全失败时才调用本地库，同时返回错误提示
        error_msg = f"AI生成失败：{str(e)}，已自动切换为本地专业标准库内容"
        local_lib = BATTERY_PROCESS_LIB if product_type == "电池包" else CHARGER_PROCESS_LIB
        local_content = local_lib.get(process_name, [])
        fallback_schemes = [
            {
                "方案名称": "本地标准库方案（AI生成失败兜底）",
                "pfmea_list": local_content
            }
        ]
        return fallback_schemes, error_msg

# 知识库解析函数（AI分析用户上传的旧PFMEA）
def parse_pfmea_knowledge(file_content, file_name):
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    
    prompt = f"""
    你是专业的PFMEA工程师，精通IATF16949和AIAG-VDA FMEA标准。
    请分析用户上传的PFMEA文件内容，提取所有符合标准的PFMEA条目，完成以下处理：
    1. 按工序名称分类，提取每个工序下的所有PFMEA条目
    2. 每个条目必须整理成标准格式，包含以下字段：
       失效模式、失效后果、失效原因、预防措施、探测措施、严重度S、频度O、探测度D、AP等级
    3. 过滤无效、重复、不符合标准的内容，补充缺失的S/O/D评分和AP等级，确保符合AIAG-VDA标准
    4. 确保所有内容贴合电池包/充电器装配场景，可直接用于PFMEA生成
    5. 返回严格的JSON格式，外层是对象，key为工序名称，value为该工序下的PFMEA条目数组，格式如下：
    {{
        "工序名称1": [
            {{
                "失效模式": "xxx",
                "失效后果": "xxx",
                "失效原因": "xxx",
                "预防措施": "xxx",
                "探测措施": "xxx",
                "严重度S": x,
                "频度O": x,
                "探测度D": x,
                "AP等级": "x"
            }}
        ]
    }}
    6. 禁止返回任何JSON以外的内容，禁止注释、解释，确保JSON可直接解析
    待解析的PFMEA文件内容：
    {file_content}
    """
    
    data = {
        "model": "doubao-pro-32k",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.3,
        "max_tokens": 8000
    }
    
    try:
        response = requests.post(API_ENDPOINT, headers=headers, json=data, timeout=60)
        response.raise_for_status()
        result = response.json()
        ai_content = result["choices"][0]["message"]["content"]
        
        # 格式容错处理
        ai_content = ai_content.strip()
        if ai_content.startswith("```json"):
            ai_content = ai_content[7:]
        if ai_content.endswith("```"):
            ai_content = ai_content[:-3]
        ai_content = ai_content.strip()
        
        # 解析JSON
        knowledge_data = json.loads(ai_content)
        return knowledge_data, None
    
    except Exception as e:
        return None, f"知识库解析失败：{str(e)}"

# Excel导出函数（兼容Excel2016，格式规范）
def export_pfmea_excel(pfmea_data, product_type, process_name):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "PFMEA"
    
    # 标题样式
    title_font = Font(name="微软雅黑", bold=True, size=14)
    header_font = Font(name="微软雅黑", bold=True, size=10)
    content_font = Font(name="微软雅黑", size=10)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # 标题
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{product_type} {process_name} 过程PFMEA"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # 标准信息
    ws.merge_cells("A2:J2")
    ws["A2"] = f"符合标准：{STANDARD}"
    ws["A2"].font = Font(name="微软雅黑", size=10)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    # 生成日期
    ws.merge_cells("A3:J3")
    ws["A3"] = f"生成日期：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = Font(name="微软雅黑", size=10)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = [
        "序号", "过程步骤/工序", "失效模式", "失效后果", "失效原因",
        "预防措施", "探测措施", "严重度S", "频度O", "探测度D", "AP等级"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border
    
    # 内容写入
    for row_num, item in enumerate(pfmea_data, 6):
        ws.cell(row=row_num, column=1, value=row_num-5).font = content_font
        ws.cell(row=row_num, column=2, value=process_name).font = content_font
        ws.cell(row=row_num, column=3, value=item["失效模式"]).font = content_font
        ws.cell(row=row_num, column=4, value=item["失效后果"]).font = content_font
        ws.cell(row=row_num, column=5, value=item["失效原因"]).font = content_font
        ws.cell(row=row_num, column=6, value=item["预防措施"]).font = content_font
        ws.cell(row=row_num, column=7, value=item["探测措施"]).font = content_font
        ws.cell(row=row_num, column=8, value=item["严重度S"]).font = content_font
        ws.cell(row=row_num, column=9, value=item["频度O"]).font = content_font
        ws.cell(row=row_num, column=10, value=item["探测度D"]).font = content_font
        ws.cell(row=row_num, column=11, value=item["AP等级"]).font = content_font
        
        # 单元格格式统一
        for col_num in range(1, 12):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = alignment
            cell.border = border
    
    # 列宽设置
    column_widths = [6, 20, 25, 30, 30, 35, 35, 8, 8, 8, 8]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+col_num)].width = width
    
    # 冻结窗格
    ws.freeze_panes = "A6"
    
    wb.save(output)
    output.seek(0)
    return output

# ===================== 3. 界面布局与主逻辑 =====================
def main():
    # 页面配置
    st.set_page_config(
        page_title=SYSTEM_NAME,
        page_icon="⚡",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # 初始化session_state
    init_session_state()
    
    # 侧边栏导航
    st.sidebar.title(SYSTEM_NAME)
    st.sidebar.markdown(f"**符合标准：** {STANDARD}")
    st.sidebar.divider()
    menu = st.sidebar.radio("功能导航", ["PFMEA智能生成", "我的知识库管理"])
    st.sidebar.divider()
    st.sidebar.markdown("**内置密钥已配置，开箱即用**")
    st.sidebar.markdown("**兼容：本地Termux | Streamlit云端 | Excel2016**")
    
    # 主页面逻辑
    if menu == "PFMEA智能生成":
        st.title("⚡ 电池包/充电器PFMEA智能生成系统")
        st.divider()
        
        # 第一步：基础设置
        st.subheader("第一步：基础参数设置")
        col1, col2, col3 = st.columns(3)
        with col1:
            product_type = st.radio("产品类型", ["电池包", "充电器"], index=0)
        with col2:
            generate_mode = st.radio("生成模式", ["本地专业标准库", "豆包AI智能生成"], index=1)
        with col3:
            # 工序选择
            process_lib = BATTERY_PROCESS_LIB if product_type == "电池包" else CHARGER_PROCESS_LIB
            # 合并用户知识库的工序
            all_process = list(process_lib.keys())
            if st.session_state.user_knowledge_base:
                user_process = list(st.session_state.user_knowledge_base.keys())
                all_process = list(set(all_process + user_process))
                all_process.sort()
            selected_process = st.selectbox("选择工序", all_process, index=0)
        
        # AI模式额外设置
        scheme_count = 3
        mix_user_knowledge = False
        if generate_mode == "豆包AI智能生成":
            st.divider()
            col4, col5 = st.columns(2)
            with col4:
                scheme_count = st.slider("AI生成方案数量", min_value=2, max_value=5, value=3, step=1)
            with col5:
                mix_user_knowledge = st.checkbox("混合我的知识库内容生成", value=False)
        
        st.divider()
        
        # 生成按钮
        generate_btn = st.button("🚀 开始生成PFMEA", type="primary", use_container_width=True)
        
        # 生成逻辑
        if generate_btn:
            # 重置状态
            st.session_state.ai_schemes = None
            st.session_state.selected_scheme = None
            st.session_state.generated_pfmea_data = None
            st.session_state.current_process = selected_process
            st.session_state.current_product = product_type
            
            with st.spinner("正在生成PFMEA内容，请稍候..."):
                if generate_mode == "本地专业标准库":
                    # 本地模式：合并标准库+用户知识库内容
                    standard_content = process_lib.get(selected_process, [])
                    user_content = st.session_state.user_knowledge_base.get(selected_process, [])
                    final_content = standard_content + user_content
                    
                    st.session_state.generated_pfmea_data = final_content
                    st.success("✅ 本地标准库PFMEA生成完成！")
                
                else:
                    # AI模式：生成多方案，彻底解决与本地重复问题
                    schemes, error_msg = generate_pfmea_ai(selected_process, product_type, scheme_count)
                    if error_msg:
                        st.warning(error_msg)
                    
                    # 混合用户知识库内容
                    if mix_user_knowledge and selected_process in st.session_state.user_knowledge_base:
                        user_content = st.session_state.user_knowledge_base[selected_process]
                        schemes.append({
                            "方案名称": "我的知识库方案",
                            "pfmea_list": user_content
                        })
                    
                    st.session_state.ai_schemes = schemes
                    st.success("✅ AI多方案生成完成！请选择您需要的方案")
        
        st.divider()
        
        # 第二步：AI方案选择（仅AI模式显示）
        if generate_mode == "豆包AI智能生成" and st.session_state.ai_schemes:
            st.subheader("第二步：选择PFMEA生成方案")
            schemes = st.session_state.ai_schemes
            scheme_names = [scheme["方案名称"] for scheme in schemes]
            
            # 方案单选
            selected_scheme_name = st.radio("请选择一个方案", scheme_names, index=0)
            selected_scheme = next(scheme for scheme in schemes if scheme["方案名称"] == selected_scheme_name)
            
            # 预览选中的方案内容
            with st.expander(f"📄 方案内容预览：{selected_scheme_name}", expanded=True):
                df_preview = pd.DataFrame(selected_scheme["pfmea_list"])
                st.dataframe(df_preview, use_container_width=True)
            
            # 确认选择按钮
            confirm_btn = st.button("✅ 确认选择此方案", type="primary", use_container_width=True)
            if confirm_btn:
                st.session_state.selected_scheme = selected_scheme
                st.session_state.generated_pfmea_data = selected_scheme["pfmea_list"]
                st.success("✅ 方案已确认！进入Excel预览环节")
        
        st.divider()
        
        # 第三步：Excel预览与调整
        if st.session_state.generated_pfmea_data:
            st.subheader("第三步：PFMEA Excel预览与调整")
            current_process = st.session_state.current_process
            current_product = st.session_state.current_product
            pfmea_data = st.session_state.generated_pfmea_data
            
            # 转换为DataFrame，支持用户在线编辑
            df = pd.DataFrame(pfmea_data)
            edited_df = st.data_editor(df, use_container_width=True, num_rows="dynamic")
            
            # 更新生成数据
            st.session_state.generated_pfmea_data = edited_df.to_dict("records")
            
            st.divider()
            
            # 第四步：Excel导出
            st.subheader("第四步：导出Excel文件")
            excel_file = export_pfmea_excel(
                st.session_state.generated_pfmea_data,
                current_product,
                current_process
            )
            
            st.download_button(
                label="📥 下载PFMEA Excel文件",
                data=excel_file,
                file_name=f"{current_product}_{current_process}_PFMEA_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
    
    # 知识库管理页面
    elif menu == "我的知识库管理":
        st.title("📚 我的PFMEA知识库管理")
        st.markdown("支持上传您现场编写的旧PFMEA文件，AI自动分析筛选入库，生成时可直接调用")
        st.divider()
        
        # 上传区域
        st.subheader("上传旧PFMEA文件入库")
        uploaded_file = st.file_uploader("请上传PFMEA Excel文件（.xlsx/.xls格式）", type=["xlsx", "xls"])
        
        if uploaded_file:
            with st.spinner("正在读取并解析文件内容，请稍候..."):
                try:
                    # 读取Excel文件所有内容
                    df_list = pd.read_excel(uploaded_file, sheet_name=None)
                    file_content = ""
                    for sheet_name, df in df_list.items():
                        file_content += f"===== 工作表：{sheet_name} =====\n"
                        file_content += df.to_string(index=False)
                        file_content += "\n\n"
                    
                    # AI解析
                    knowledge_data, error_msg = parse_pfmea_knowledge(file_content, uploaded_file.name)
                    
                    if error_msg:
                        st.error(error_msg)
                    else:
                        st.success("✅ 文件解析完成！请确认要入库的内容")
                        
                        # 展示解析结果，让用户确认
                        with st.expander("📄 解析结果预览", expanded=True):
                            for process_name, pfmea_list in knowledge_data.items():
                                st.markdown(f"**工序名称：{process_name}**")
                                st.dataframe(pd.DataFrame(pfmea_list), use_container_width=True)
                                st.divider()
                        
                        # 入库按钮
                        storage_btn = st.button("✅ 确认入库", type="primary", use_container_width=True)
                        if storage_btn:
                            # 合并到用户知识库
                            for process_name, pfmea_list in knowledge_data.items():
                                if process_name in st.session_state.user_knowledge_base:
                                    # 去重合并
                                    existing_list = st.session_state.user_knowledge_base[process_name]
                                    existing_keys = set(f"{item['失效模式']}_{item['失效原因']}" for item in existing_list)
                                    for item in pfmea_list:
                                        item_key = f"{item['失效模式']}_{item['失效原因']}"
                                        if item_key not in existing_keys:
                                            existing_list.append(item)
                                    st.session_state.user_knowledge_base[process_name] = existing_list
                                else:
                                    st.session_state.user_knowledge_base[process_name] = pfmea_list
                            
                            st.success("✅ 知识库入库完成！生成PFMEA时可直接调用")
                            st.rerun()
                
                except Exception as e:
                    st.error(f"文件读取失败：{str(e)}，请检查文件格式是否正确")
        
        st.divider()
        
        # 知识库内容管理
        st.subheader("我的知识库内容")
        user_kb = st.session_state.user_knowledge_base
        
        if not user_kb:
            st.info("您的知识库暂无内容，请上传PFMEA文件入库")
        else:
            # 遍历所有工序内容
            for process_name, pfmea_list in user_kb.items():
                with st.expander(f"📦 工序：{process_name}（共{len(pfmea_list)}条PFMEA）", expanded=False):
                    # 可编辑的DataFrame
                    edited_df = st.data_editor(
                        pd.DataFrame(pfmea_list),
                        use_container_width=True,
                        num_rows="dynamic",
                        key=f"edit_{process_name}"
                    )
                    
                    # 更新按钮
                    col_update, col_delete = st.columns(2)
                    with col_update:
                        if st.button("✅ 更新内容", key=f"update_{process_name}", use_container_width=True):
                            st.session_state.user_knowledge_base[process_name] = edited_df.to_dict("records")
                            st.success("内容更新成功！")
                            st.rerun()
                    
                    with col_delete:
                        if st.button("🗑️ 删除此工序", key=f"delete_{process_name}", use_container_width=True, type="secondary"):
                            del st.session_state.user_knowledge_base[process_name]
                            st.success("工序删除成功！")
                            st.rerun()
                
                st.divider()
            
            # 知识库导出/导入
            st.subheader("知识库备份与恢复")
            col_export, col_import = st.columns(2)
            with col_export:
                # 导出知识库为JSON
                kb_json = json.dumps(user_kb, ensure_ascii=False, indent=2)
                st.download_button(
                    label="📤 导出知识库备份文件",
                    data=kb_json,
                    file_name=f"PFMEA知识库备份_{datetime.now().strftime('%Y%m%d%H%M%S')}.json",
                    mime="application/json",
                    use_container_width=True
                )
            
            with col_import:
                # 导入知识库
                import_file = st.file_uploader("导入知识库备份文件", type=["json"], label_visibility="collapsed")
                if import_file:
                    try:
                        import_data = json.load(import_file)
                        if st.button("✅ 确认导入", use_container_width=True):
                            # 合并导入的知识库
                            for process_name, pfmea_list in import_data.items():
                                if process_name in st.session_state.user_knowledge_base:
                                    existing_list = st.session_state.user_knowledge_base[process_name]
                                    existing_keys = set(f"{item['失效模式']}_{item['失效原因']}" for item in existing_list)
                                    for item in pfmea_list:
                                        item_key = f"{item['失效模式']}_{item['失效原因']}"
                                        if item_key not in existing_keys:
                                            existing_list.append(item)
                                    st.session_state.user_knowledge_base[process_name] = existing_list
                                else:
                                    st.session_state.user_knowledge_base[process_name] = pfmea_list
                            st.success("✅ 知识库导入完成！")
                            st.rerun()
                    except Exception as e:
                        st.error(f"导入失败：{str(e)}，请检查备份文件格式是否正确")

# 程序入口
if __name__ == "__main__":
    main()
